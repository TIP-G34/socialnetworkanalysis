import pandas as pd
from kuzu import Connection
import openpyxl
import uuid
import os
import time

class Storage:

    def __init__(self, k_client: Connection):
        # Initialize database and create schema.
        self.k_client = k_client
        

    def clean_data(self, df):
        # reads the data provided in pandas dataframe and cleans missing values.
        df.replace('N/A', None, inplace=True)         # Replace missing values with '0'
        df.replace('NA', None, inplace=True)
        non_none_columns = [col for col in df.columns if col is not None]
        # Create a new DataFrame containing only columns with non-None names
        df = df[non_none_columns]
        return df
    
    def load_ids(self, path1, path2):
        # path 1 is path to affiliation csv sheet
        # path 2 is path to participants csv sheet
        df1 = self.clean_ingested_data(path1)        # affiliations
        df2 = self.clean_ingested_data(path2)        # participants

        df1 = df1[['ID', 'Title']]
        df2 = df2['Participant-ID']
        
        df1.to_csv("../Datasets/participants", index=False)     # Save cleaned csv file to Datasets folder
        df2.to_csv("../Datasets/affiliations", index=False)

    def load_data(self,file_type,data_type,file_path,helper):
        status = False
        message = ""
        if file_type == 'csv':
            return self.process_csv(data_type,file_path)
        elif file_type == 'excel':
            return self.process_excel(file_path,helper)
        else:
            status = False
            message = "Invalid file type"
        
        return {
                'status': status,
                'message': message
                    }


    def process_excel(self, file_path,helper):
        try:
            workbook = openpyxl.load_workbook(file_path,data_only=True)
        except Exception as e:
            return {
                'status': False,
                'message': str(e)
            }

        all_data = {}

        try:
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]            
                sheet_data = []

                is_header_row = True
                for row in sheet.iter_rows(values_only=True):
                    if is_header_row:
                        header_row = row
                        is_header_row = False
                        continue

                    sheet_data.append(dict(zip(header_row, row)))
                all_data[sheet_name] = sheet_data
        except Exception as e:
            return {
                'status': False,
                'message': str(e)
            }
        
        try:
            for key, value in all_data.items():
                if "net_" not in key and "data_dictionary" not in key:
                    cleaned_data = self.clean_data(pd.DataFrame(value))     
                    create_table_response = self.create_tables(helper[key])  
                    if create_table_response['status'] == False:
                        return create_table_response
                    
                    if not os.path.exists("../tmp"):
                        os.makedirs("../tmp")
                    
                    tmp_file_path = f"""../tmp/{str(uuid.uuid4()).replace("-","")}.csv"""
                    cleaned_data.to_csv(tmp_file_path, index=False,header=False)
                    self.ingest_data(helper[key],tmp_file_path)
                    # os.remove(tmp_file_path)

                    continue
        except Exception as e:
            return {
                'status': False,
                'message': str(e)
            }
        
        try:
            for key, value in all_data.items():
                if "net_" in key:
                    cleaned_data = self.clean_data(pd.DataFrame(value))     
                    create_rel_response = self.create_relationships(helper[key])  
                    if create_rel_response['status'] == False:
                        return create_rel_response
                    
                    if not os.path.exists("../tmp"):
                        os.makedirs("../tmp")
                    
                    tmp_file_path = f"""../tmp/{str(uuid.uuid4()).replace("-","")}.csv"""
                    cleaned_data.to_csv(tmp_file_path, index=False,header=False)
                    self.ingest_relationships(helper[key],tmp_file_path)
                    # os.remove(tmp_file_path)

                    continue
        except Exception as e:
            return {
                'status': False,
                'message': str(e)
            }


    # 

    def create_tables(self, helper):
        columns = ""
        print(helper)
        for key, value in helper["columns"].items():
            columns += f"{key} {value},"

        columns += f"""PRIMARY KEY ({helper["primary_key"]})"""
        try:
            self.k_client.execute(f"""CREATE NODE TABLE {helper["table_name"]}(
                {columns}
            )""")
        except Exception as e:
            if "already exists" in str(e):
                return {
                    'status': True,
                    'message': "Table already exists"
                }
        
        return {
            'status': True,
            'message': "Table created successfully"
        }
    
    def create_relationships(self, helper):
        # CREATE REL TABLE LivesIn(FROM User TO City)
        try:
            self.k_client.execute(f"""CREATE REL TABLE {helper["relationship_name"]}(
                FROM {helper["source"]} TO {helper["target"]}
            )""")
        except Exception as e:
            if "already exists" in str(e):
                return {
                    'status': True,
                    'message': "Relationship already exists"
                }
        
        return {
            'status': True,
            'message': "Relationship created successfully"
        }
    
    def ingest_relationships(self,helper,tmp_csv_path):
        try:
            self.k_client.execute(f"""COPY {helper["relationship_name"]} FROM "{tmp_csv_path}" """)
            return {
                'status': True,
                'message': "Relationsip ingested successfully"
            }
        except Exception as e:
            print(str(e))
            if "COPY commands can only" in str(e):
                return {
                    'status': True,
                    'message': "Relationship already exists"
                }
    

        # print(self.k_client.execute("""
        #     CALL db.schema()
        #     YIELD labels
        #     WHERE 'YourLabelName' IN labels
        #     RETURN 'YourLabelName' AS labelExists
        #     """))

    def ingest_data(self,helper,tmp_csv_path):
        try:
            self.k_client.execute(f"""COPY {helper["table_name"]} FROM "{tmp_csv_path}" """)
            return {
                'status': True,
                'message': "Data ingested successfully"
            }
        except Exception as e:
            print(str(e))
            if "COPY commands can only" in str(e):
                return {
                    'status': True,
                    'message': "Data already exists"
                }
        

    
    def upload_data(self):
        # Upload data to DB
        # Save data to KuzuDB

        self.k_client.execute('COPY friends FROM "../Datasets/friends.csv"')
        self.k_client.execute('COPY influential FROM "../Datasets/influential.csv"')
        self.k_client.execute('COPY feedback FROM "../Datasets/feedback.csv"')
        self.k_client.execute('COPY moretime FROM "../Datasets/moretime.csv"')
        self.k_client.execute('COPY advice FROM "../Datasets/advice.csv"')
        self.k_client.execute('COPY disrespect FROM "../Datasets/disrespect.csv"')
        self.k_client.execute('COPY schoolactivity FROM "../Datasets/activity.csv"')